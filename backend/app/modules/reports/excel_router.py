from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from datetime import datetime

from app.core.dependencies import get_db, get_current_user
from app.modules.patients.models import Patient, Gender, BloodType
from app.modules.appointments.models import Appointment, AppointmentStatus
from app.modules.billing.models import Invoice

router = APIRouter(prefix="/excel", tags=["Excel"])


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0A3D6B", end_color="0A3D6B", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


@router.get("/export/patients")
def export_patients(db: Session = Depends(get_db), _=Depends(get_current_user)):
    patients = db.query(Patient).filter(Patient.is_deleted == False).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    headers = ["ID", "Tipo Documento", "Numero Documento", "Nombres", "Apellidos",
               "Fecha Nacimiento", "Genero", "Tipo Sangre", "Email", "Telefono",
               "Direccion", "Ciudad", "Alergias", "Notas", "Fecha Registro"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[cell.column_letter].width = 18

    for row, p in enumerate(patients, 2):
        ws.cell(row=row, column=1, value=str(p.id))
        ws.cell(row=row, column=2, value=p.document_type)
        ws.cell(row=row, column=3, value=p.document_number)
        ws.cell(row=row, column=4, value=p.first_name)
        ws.cell(row=row, column=5, value=p.last_name)
        ws.cell(row=row, column=6, value=str(p.date_of_birth) if p.date_of_birth else "")
        ws.cell(row=row, column=7, value=p.gender.value if p.gender else "")
        ws.cell(row=row, column=8, value=p.blood_type.value if p.blood_type else "")
        ws.cell(row=row, column=9, value=p.email or "")
        ws.cell(row=row, column=10, value=p.phone or "")
        ws.cell(row=row, column=11, value=p.address or "")
        ws.cell(row=row, column=12, value=p.city or "")
        ws.cell(row=row, column=13, value=p.allergies or "")
        ws.cell(row=row, column=14, value=p.notes or "")
        ws.cell(row=row, column=15, value=str(p.created_at) if p.created_at else "")
        if row % 2 == 0:
            for col in range(1, 16):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"pacientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/appointments")
def export_appointments(db: Session = Depends(get_db), _=Depends(get_current_user)):
    appointments = db.query(Appointment).filter(Appointment.is_deleted == False).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Citas"

    headers = ["ID", "Paciente ID", "Doctor ID", "Fecha y Hora", "Duracion (min)", "Estado", "Motivo", "Notas"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[cell.column_letter].width = 20

    for row, a in enumerate(appointments, 2):
        ws.cell(row=row, column=1, value=str(a.id))
        ws.cell(row=row, column=2, value=str(a.patient_id))
        ws.cell(row=row, column=3, value=str(a.doctor_id))
        ws.cell(row=row, column=4, value=str(a.scheduled_at) if a.scheduled_at else "")
        ws.cell(row=row, column=5, value=a.duration_minutes)
        ws.cell(row=row, column=6, value=a.status.value if a.status else "")
        ws.cell(row=row, column=7, value=a.reason or "")
        ws.cell(row=row, column=8, value=a.notes or "")
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"citas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/invoices")
def export_invoices(db: Session = Depends(get_db), _=Depends(get_current_user)):
    invoices = db.query(Invoice).filter(Invoice.is_deleted == False).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = ["ID", "Numero", "Paciente ID", "Estado", "Subtotal", "IVA", "Total", "Notas", "Fecha"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[cell.column_letter].width = 18

    for row, inv in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=str(inv.id))
        ws.cell(row=row, column=2, value=inv.number)
        ws.cell(row=row, column=3, value=str(inv.patient_id))
        ws.cell(row=row, column=4, value=inv.status.value if inv.status else "")
        ws.cell(row=row, column=5, value=float(inv.subtotal or 0))
        ws.cell(row=row, column=6, value=float(inv.tax or 0))
        ws.cell(row=row, column=7, value=float(inv.total or 0))
        ws.cell(row=row, column=8, value=inv.notes or "")
        ws.cell(row=row, column=9, value=str(inv.created_at) if inv.created_at else "")
        if row % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"facturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/template/patients")
def download_patient_template(_=Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Pacientes"

    headers = ["tipo_documento", "numero_documento", "nombres", "apellidos",
               "fecha_nacimiento", "genero", "tipo_sangre", "email",
               "telefono", "direccion", "ciudad", "alergias", "notas"]

    notes = ["cedula/pasaporte/ruc", "Ej: 0102030405", "Juan Carlos", "Perez Lopez",
             "YYYY-MM-DD", "male/female/other", "A+/A-/B+/B-/AB+/AB-/O+/O-",
             "Opcional", "Opcional", "Opcional", "Opcional", "Opcional", "Opcional"]

    for col, (header, note) in enumerate(zip(headers, notes), 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[cell.column_letter].width = 20
        ws.cell(row=2, column=col, value=note).font = Font(italic=True, color="9CA3AF")

    ws.cell(row=3, column=1, value="cedula")
    ws.cell(row=3, column=2, value="0102030405")
    ws.cell(row=3, column=3, value="Juan")
    ws.cell(row=3, column=4, value="Perez")
    ws.cell(row=3, column=5, value="1990-05-15")
    ws.cell(row=3, column=6, value="male")
    ws.cell(row=3, column=7, value="O+")
    ws.cell(row=3, column=8, value="juan@email.com")
    ws.cell(row=3, column=9, value="0991234567")
    ws.cell(row=3, column=10, value="Av. Principal 123")
    ws.cell(row=3, column=11, value="Quito")
    ws.cell(row=3, column=12, value="")
    ws.cell(row=3, column=13, value="")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_pacientes.xlsx"})


@router.post("/import/patients")
async def import_patients(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls)")

    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content), skiprows=1)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")

    created = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            doc_number = str(row.get("numero_documento", "")).strip()
            if not doc_number or doc_number == "nan":
                continue

            existing = db.query(Patient).filter(Patient.document_number == doc_number, Patient.is_deleted == False).first()
            if existing:
                errors.append(f"Fila {idx+3}: Documento {doc_number} ya existe")
                continue

            fecha = row.get("fecha_nacimiento", "")
            if pd.isna(fecha):
                errors.append(f"Fila {idx+3}: Fecha de nacimiento requerida")
                continue

            from datetime import date
            if isinstance(fecha, str):
                fecha_parsed = datetime.strptime(fecha.strip(), "%Y-%m-%d").date()
            else:
                fecha_parsed = fecha.date() if hasattr(fecha, 'date') else date.fromisoformat(str(fecha)[:10])

            genero_val = str(row.get("genero", "other")).strip().lower()
            if genero_val not in ["male", "female", "other"]:
                genero_val = "other"

            blood_val = str(row.get("tipo_sangre", "")).strip()
            blood_type = None
            valid_bloods = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
            if blood_val in valid_bloods:
                blood_type = BloodType(blood_val)

            patient = Patient(
                document_type=str(row.get("tipo_documento", "cedula")).strip(),
                document_number=doc_number,
                first_name=str(row.get("nombres", "")).strip(),
                last_name=str(row.get("apellidos", "")).strip(),
                date_of_birth=fecha_parsed,
                gender=Gender(genero_val),
                blood_type=blood_type,
                email=str(row.get("email", "")).strip() if not pd.isna(row.get("email", float("nan"))) else None,
                phone=str(row.get("telefono", "")).strip() if not pd.isna(row.get("telefono", float("nan"))) else None,
                address=str(row.get("direccion", "")).strip() if not pd.isna(row.get("direccion", float("nan"))) else None,
                city=str(row.get("ciudad", "")).strip() if not pd.isna(row.get("ciudad", float("nan"))) else None,
                allergies=str(row.get("alergias", "")).strip() if not pd.isna(row.get("alergias", float("nan"))) else None,
                notes=str(row.get("notas", "")).strip() if not pd.isna(row.get("notas", float("nan"))) else None,
            )
            db.add(patient)
            created += 1
        except Exception as e:
            errors.append(f"Fila {idx+3}: {str(e)}")

    db.commit()
    return {"created": created, "errors": errors, "message": f"{created} pacientes importados correctamente"}
